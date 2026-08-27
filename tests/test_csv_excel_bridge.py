from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.csv_excel_bridge import (
    BridgeError,
    UTF8_BOM,
    compare_csv_logical,
    export_csv_to_xlsx,
    import_xlsx_to_csv,
    read_csv_records,
    serialize_csv,
    validate_csv,
)


def write_csv(path: Path, rows: list[list[str]], bom: bool = False) -> Path:
    raw = serialize_csv(rows)
    path.write_bytes((UTF8_BOM if bom else b"") + raw)
    return path


def sample_rows() -> list[list[str]]:
    return [
        ["identifier", "description", "date_looking", "empty"],
        ["000123", "日本語", "2026-08-27", ""],
        ["12345678901234567890", 'comma, quote " and\nembedded newline', "01/02/2026", "  significant  "],
    ]


def export_sample(tmp_path: Path, rows: list[list[str]] | None = None) -> tuple[Path, Path]:
    csv_path = write_csv(tmp_path / "source.csv", rows or sample_rows())
    xlsx_path = tmp_path / "review.xlsx"
    export_csv_to_xlsx(csv_path, xlsx_path)
    return csv_path, xlsx_path


def test_roundtrip_is_byte_identical(tmp_path: Path) -> None:
    source, review = export_sample(tmp_path)
    output = tmp_path / "output.csv"
    import_xlsx_to_csv(review, output, baseline=source)
    assert output.read_bytes() == source.read_bytes()


def test_bom_input_is_accepted_and_output_has_no_bom(tmp_path: Path) -> None:
    source = write_csv(tmp_path / "bom.csv", sample_rows(), bom=True)
    review = tmp_path / "review.xlsx"
    output = tmp_path / "output.csv"
    export_csv_to_xlsx(source, review, encoding="auto-utf8")
    import_xlsx_to_csv(review, output, baseline=source)
    assert not output.read_bytes().startswith(UTF8_BOM)


@pytest.mark.parametrize(
    "value",
    [
        "日本語",
        "000123",
        "2026-08-27",
        "12345678901234567890",
        "",
        "comma,value",
        'quote " value',
        "embedded\nnewline",
        "  significant  ",
    ],
)
def test_lexical_values_remain_strings(tmp_path: Path, value: str) -> None:
    source, review = export_sample(tmp_path, [["value"], [value]])
    workbook = load_workbook(review, data_only=False)
    cell = workbook["CSV"]["A2"]
    assert cell.value == (value if value != "" else None)
    assert cell.number_format == "@"
    output = tmp_path / "output.csv"
    import_xlsx_to_csv(review, output, baseline=source)
    assert read_csv_records(output).records == [["value"], [value]]


@pytest.mark.parametrize(
    ("value", "data_type"),
    [("=1+1", "f"), (123, "n"), (date(2026, 8, 27), "d"), (True, "b")],
)
def test_nontext_cells_are_rejected_and_leave_no_output(
    tmp_path: Path, value: object, data_type: str
) -> None:
    source, review = export_sample(tmp_path, [["value"], ["text"]])
    workbook = load_workbook(review)
    workbook["CSV"]["A2"] = value
    workbook.save(review)
    output = tmp_path / "rejected.csv"
    with pytest.raises(BridgeError):
        import_xlsx_to_csv(review, output, baseline=source)
    assert not output.exists()


def test_unexpected_visible_worksheet_rejected(tmp_path: Path) -> None:
    source, review = export_sample(tmp_path)
    workbook = load_workbook(review)
    workbook.create_sheet("Unexpected")
    workbook.save(review)
    with pytest.raises(BridgeError):
        import_xlsx_to_csv(review, tmp_path / "output.csv", baseline=source)


def test_hidden_metadata_worksheet_allowed(tmp_path: Path) -> None:
    source, review = export_sample(tmp_path)
    workbook = load_workbook(review)
    metadata = workbook.create_sheet("__UADC_BRIDGE__")
    metadata.sheet_state = "hidden"
    workbook.save(review)
    import_xlsx_to_csv(review, tmp_path / "output.csv", baseline=source)


def test_header_change_rejected(tmp_path: Path) -> None:
    source, review = export_sample(tmp_path)
    workbook = load_workbook(review)
    workbook["CSV"]["A1"] = "changed"
    workbook.save(review)
    with pytest.raises(BridgeError):
        import_xlsx_to_csv(review, tmp_path / "output.csv", baseline=source)


def test_row_count_change_rejected(tmp_path: Path) -> None:
    source, review = export_sample(tmp_path)
    workbook = load_workbook(review)
    workbook["CSV"].delete_rows(3)
    workbook.save(review)
    with pytest.raises(BridgeError):
        import_xlsx_to_csv(review, tmp_path / "output.csv", baseline=source)


def test_column_count_change_rejected(tmp_path: Path) -> None:
    source, review = export_sample(tmp_path)
    workbook = load_workbook(review)
    workbook["CSV"].delete_cols(4)
    workbook.save(review)
    with pytest.raises(BridgeError):
        import_xlsx_to_csv(review, tmp_path / "output.csv", baseline=source)


def test_intended_text_edit_manifest_hashes_values(tmp_path: Path) -> None:
    source, review = export_sample(tmp_path)
    workbook = load_workbook(review)
    workbook["CSV"]["B2"] = "公開用編集"
    workbook.save(review)
    output = tmp_path / "output.csv"
    manifest = tmp_path / "manifest.json"
    result = import_xlsx_to_csv(review, output, baseline=source, manifest_path=manifest)
    assert result["changed_cell_count"] == 1
    text = manifest.read_text(encoding="utf-8")
    assert "公開用編集" not in text
    assert result["changed_cells"][0]["row"] == 2


def test_output_has_lf_and_exactly_one_final_lf(tmp_path: Path) -> None:
    source, review = export_sample(tmp_path)
    output = tmp_path / "output.csv"
    import_xlsx_to_csv(review, output, baseline=source)
    raw = output.read_bytes()
    assert b"\r" not in raw
    assert raw.endswith(b"\n") and not raw.endswith(b"\n\n")


def test_import_preserves_crlf_record_delimiters_and_quoted_newlines(tmp_path: Path) -> None:
    source = tmp_path / "source.csv"
    source.write_bytes(b'identifier,description\r\n000123,"first\r\nsecond"\r\n')
    review = tmp_path / "review.xlsx"
    output = tmp_path / "output.csv"
    export_csv_to_xlsx(source, review)
    import_xlsx_to_csv(review, output, baseline=source, output_eol="preserve")
    assert output.read_bytes() == source.read_bytes()


@pytest.mark.parametrize(
    ("mode", "record_delimiter"), [("lf", b"\n"), ("crlf", b"\r\n")]
)
def test_import_allows_explicit_record_eol_without_changing_quoted_value(
    tmp_path: Path, mode: str, record_delimiter: bytes
) -> None:
    source = write_csv(
        tmp_path / "source.csv",
        [["identifier", "description"], ["000123", "first\r\nsecond"]],
    )
    review = tmp_path / "review.xlsx"
    output = tmp_path / "output.csv"
    export_csv_to_xlsx(source, review)
    import_xlsx_to_csv(review, output, baseline=source, output_eol=mode)
    raw = output.read_bytes()
    assert raw.startswith(b"identifier,description" + record_delimiter)
    assert b"first\r\nsecond" in raw
    assert read_csv_records(output).records == read_csv_records(source).records


def test_compare_treats_only_quoted_newline_style_as_logically_identical(
    tmp_path: Path,
) -> None:
    left = tmp_path / "left.csv"
    right = tmp_path / "right.csv"
    left.write_bytes(b'h\n"first\r\nsecond"\n')
    right.write_bytes(b'h\n"first\nsecond"\n')
    assert compare_csv_logical(left, right)[0] == "LOGICALLY_IDENTICAL_BYTES_DIFFER"


def test_compare_byte_identical(tmp_path: Path) -> None:
    left = write_csv(tmp_path / "left.csv", [["a"], ["b"]])
    right = tmp_path / "right.csv"
    right.write_bytes(left.read_bytes())
    assert compare_csv_logical(left, right)[0] == "BYTE_IDENTICAL"


def test_compare_logically_identical_bytes_differ(tmp_path: Path) -> None:
    left = write_csv(tmp_path / "left.csv", [["a"], ["b"]])
    right = tmp_path / "right.csv"
    right.write_bytes(UTF8_BOM + left.read_bytes())
    assert compare_csv_logical(left, right)[0] == "LOGICALLY_IDENTICAL_BYTES_DIFFER"


def test_compare_logical_difference(tmp_path: Path) -> None:
    left = write_csv(tmp_path / "left.csv", [["a"], ["b"]])
    right = write_csv(tmp_path / "right.csv", [["a"], ["c"]])
    assert compare_csv_logical(left, right)[0] == "LOGICAL_DIFFERENCE"


def test_compare_parse_error(tmp_path: Path) -> None:
    left = write_csv(tmp_path / "left.csv", [["a"]])
    right = tmp_path / "right.csv"
    right.write_bytes(b"\xff")
    assert compare_csv_logical(left, right)[0] == "PARSE_ERROR"


def make_template(path: Path) -> Path:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "CSV"
    ws["A1"] = "STALE_HEADER"
    ws["A2"] = "STALE_VALUE"
    ws["B2"] = "STALE_BLANK_REPLACEMENT"
    ws["A3"] = "=1+1"
    ws.column_dimensions["A"].width = 27
    ws.column_dimensions["B"].hidden = True
    ws.row_dimensions[2].height = 31
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = "A1:B3"
    ws["A1"].font = Font(bold=True, color="FF0000")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    ws["A1"].border = Border(bottom=Side(style="thin", color="000000"))
    ws["A1"].alignment = Alignment(horizontal="center", wrap_text=True)
    workbook.save(path)
    return path


def test_template_restores_formatting_and_csv_values_are_authoritative(tmp_path: Path) -> None:
    source = write_csv(tmp_path / "source.csv", [["h1", "h2"], ["CSV_VALUE", ""]])
    template = make_template(tmp_path / "template.xlsx")
    output = tmp_path / "output.xlsx"
    manifest = tmp_path / "manifest.json"
    result = export_csv_to_xlsx(source, output, template=template, manifest_path=manifest)
    workbook = load_workbook(output, data_only=False)
    ws = workbook["CSV"]
    assert ws["A2"].value == "CSV_VALUE"
    assert ws["B2"].value is None
    assert ws["A3"].value is None
    assert ws.column_dimensions["A"].width == 27
    assert ws.column_dimensions["B"].hidden is True
    assert ws.row_dimensions[2].height == 31
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:B2"
    assert ws["A1"].font.bold is True
    assert ws["A1"].fill.fgColor.rgb == "00FFFF00"
    assert ws["A1"].border.bottom.style == "thin"
    assert ws["A1"].alignment.horizontal == "center"
    assert ws["A2"].number_format == "@"
    assert result["template_sha256"]
    assert result["formatting_recovery"] is True


def test_template_formula_inside_data_rectangle_is_removed(tmp_path: Path) -> None:
    source = write_csv(tmp_path / "source.csv", [["h"], ["text"], ["third"]])
    template = make_template(tmp_path / "template.xlsx")
    output = tmp_path / "output.xlsx"
    export_csv_to_xlsx(source, output, template=template)
    assert load_workbook(output, data_only=False)["CSV"]["A3"].value == "third"


def test_additional_rows_receive_text_format(tmp_path: Path) -> None:
    source = write_csv(tmp_path / "source.csv", [["h"], ["one"], ["two"], ["three"]])
    template = make_template(tmp_path / "template.xlsx")
    output = tmp_path / "output.xlsx"
    export_csv_to_xlsx(source, output, template=template)
    assert load_workbook(output)["CSV"]["A4"].number_format == "@"


def test_template_with_incompatible_merge_rejected(tmp_path: Path) -> None:
    source = write_csv(tmp_path / "source.csv", [["h1", "h2"], ["a", "b"]])
    workbook = Workbook()
    workbook.active.title = "CSV"
    workbook.active.merge_cells("A1:B1")
    template = tmp_path / "template.xlsx"
    workbook.save(template)
    with pytest.raises(BridgeError):
        export_csv_to_xlsx(source, tmp_path / "output.xlsx", template=template)


def test_template_missing_data_sheet_rejected(tmp_path: Path) -> None:
    source = write_csv(tmp_path / "source.csv", [["h"], ["a"]])
    workbook = Workbook()
    workbook.active.title = "Other"
    template = tmp_path / "template.xlsx"
    workbook.save(template)
    with pytest.raises(BridgeError):
        export_csv_to_xlsx(source, tmp_path / "output.xlsx", template=template)


@pytest.mark.parametrize("lexical", ["2026-08-27", "2026/08/27", "2026.08.27"])
def test_declared_date_text_normalizes_to_iso(tmp_path: Path, lexical: str) -> None:
    source = write_csv(tmp_path / "source.csv", [["date"], [lexical]])
    review = tmp_path / "review.xlsx"
    export_csv_to_xlsx(source, review, date_columns=["date"])
    output = tmp_path / "output.csv"
    import_xlsx_to_csv(review, output, baseline=source, date_columns=["date"])
    assert read_csv_records(output).records[1][0] == "2026-08-27"


def test_true_excel_date_in_declared_column_is_accepted(tmp_path: Path) -> None:
    source, review = export_sample(tmp_path, [["date"], ["2026-08-27"]])
    workbook = load_workbook(review)
    workbook["CSV"]["A2"] = date(2026, 8, 27)
    workbook.save(review)
    output = tmp_path / "output.csv"
    import_xlsx_to_csv(review, output, baseline=source, date_columns=["date"])
    assert read_csv_records(output).records[1][0] == "2026-08-27"


def test_true_excel_date_outside_declared_column_is_rejected(tmp_path: Path) -> None:
    source, review = export_sample(tmp_path, [["value"], ["text"]])
    workbook = load_workbook(review)
    workbook["CSV"]["A2"] = date(2026, 8, 27)
    workbook.save(review)
    with pytest.raises(BridgeError):
        import_xlsx_to_csv(review, tmp_path / "output.csv", baseline=source)


def test_ambiguous_date_text_is_rejected(tmp_path: Path) -> None:
    source, review = export_sample(tmp_path, [["date"], ["01/02/2026"]])
    with pytest.raises(BridgeError):
        import_xlsx_to_csv(review, tmp_path / "output.csv", baseline=source, date_columns=["date"])


def test_explicit_date_format_removes_ambiguity(tmp_path: Path) -> None:
    source, review = export_sample(tmp_path, [["date"], ["01/02/2026"]])
    output = tmp_path / "output.csv"
    import_xlsx_to_csv(
        review,
        output,
        baseline=source,
        date_columns=["date"],
        date_formats=["%d/%m/%Y"],
    )
    assert read_csv_records(output).records[1][0] == "2026-02-01"


def test_datetime_with_time_is_rejected(tmp_path: Path) -> None:
    source, review = export_sample(tmp_path, [["date"], ["2026-08-27"]])
    workbook = load_workbook(review)
    workbook["CSV"]["A2"] = datetime(2026, 8, 27, 12, 30)
    workbook.save(review)
    with pytest.raises(BridgeError):
        import_xlsx_to_csv(review, tmp_path / "output.csv", baseline=source, date_columns=["date"])


def test_empty_declared_date_cell_remains_empty(tmp_path: Path) -> None:
    source, review = export_sample(tmp_path, [["date"], [""]])
    output = tmp_path / "output.csv"
    import_xlsx_to_csv(review, output, baseline=source, date_columns=["date"])
    assert read_csv_records(output).records == [["date"], [""]]


def test_validate_canonical_passes(tmp_path: Path) -> None:
    source = write_csv(tmp_path / "source.csv", [["a", "b"], ["1", "2"]])
    assert validate_csv(source, canonical=True, expected_columns=2)["result"] == "PASS"


def test_validate_canonical_rejects_bom(tmp_path: Path) -> None:
    source = write_csv(tmp_path / "source.csv", [["a"], ["b"]], bom=True)
    with pytest.raises(BridgeError):
        validate_csv(source, canonical=True)


def test_strict_utf8_export_rejects_bom(tmp_path: Path) -> None:
    source = write_csv(tmp_path / "source.csv", [["a"], ["b"]], bom=True)
    with pytest.raises(BridgeError):
        export_csv_to_xlsx(source, tmp_path / "output.xlsx", encoding="utf-8")


def test_declared_date_column_must_exist(tmp_path: Path) -> None:
    source = write_csv(tmp_path / "source.csv", [["other"], ["text"]])
    with pytest.raises(BridgeError):
        export_csv_to_xlsx(source, tmp_path / "output.xlsx", date_columns=["date"])
