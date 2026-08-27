"""Controlled bridge between Canonical CSV and review-only XLSX files.

Canonical CSV values are authoritative.  XLSX templates contribute formatting
only; they never contribute data values, formulas, or cached results.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import tempfile
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


FORMAT_VERSION = "1.0"
UTF8_BOM = b"\xef\xbb\xbf"
DEFAULT_DATE_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d")


class BridgeError(Exception):
    """Expected validation or I/O failure with a stable CLI exit code."""

    def __init__(self, message: str, code: int = 7) -> None:
        super().__init__(message)
        self.code = code


@dataclass(frozen=True)
class CsvDocument:
    path: Path
    records: list[list[str]]
    raw: bytes
    encoding: str
    bom: bool

    @property
    def record_count(self) -> int:
        return len(self.records)

    @property
    def column_count(self) -> int:
        return max((len(row) for row in self.records), default=0)


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest().upper()


def sha256_file(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def value_sha256(value: str) -> str:
    return sha256_bytes(value.encode("utf-8"))


def _eol_label(raw: bytes) -> str:
    if b"\r\n" in raw:
        return "CRLF_OR_MIXED"
    if b"\r" in raw:
        return "CR_OR_MIXED"
    return "LF"


def _decode_utf8(raw: bytes, mode: str) -> tuple[str, bool, str]:
    bom = raw.startswith(UTF8_BOM)
    if mode == "utf-8" and bom:
        raise BridgeError("UTF-8 BOM is forbidden in strict utf-8 mode", 3)
    if mode not in {"utf-8", "utf-8-sig", "auto-utf8"}:
        raise BridgeError(f"Unsupported encoding mode: {mode}", 2)
    payload = raw[len(UTF8_BOM) :] if bom else raw
    try:
        return payload.decode("utf-8", errors="strict"), bom, (
            "utf-8-sig" if bom else "utf-8"
        )
    except UnicodeDecodeError as exc:
        raise BridgeError(f"Input is not valid UTF-8 at byte {exc.start}", 3) from exc


def read_csv_records(path: Path, encoding: str = "auto-utf8") -> CsvDocument:
    raw = path.read_bytes()
    if b"\x00" in raw:
        raise BridgeError("CSV contains a NUL byte", 4)
    text, bom, actual_encoding = _decode_utf8(raw, encoding)
    try:
        records = [list(row) for row in csv.reader(io.StringIO(text, newline=""))]
    except csv.Error as exc:
        raise BridgeError(f"CSV parse error: {exc}", 4) from exc
    return CsvDocument(path, records, raw, actual_encoding, bom)


def _record_eol(raw: bytes) -> str:
    """Report the record-delimiter style without inspecting quoted newlines as delimiters."""

    text = raw.decode("utf-8-sig", errors="strict")
    delimiters: list[str] = []
    quoted = False
    index = 0
    while index < len(text):
        char = text[index]
        if char == '"':
            if quoted and index + 1 < len(text) and text[index + 1] == '"':
                index += 2
                continue
            quoted = not quoted
        elif not quoted and char == "\r":
            if index + 1 < len(text) and text[index + 1] == "\n":
                delimiters.append("CRLF")
                index += 2
                continue
            delimiters.append("CR")
        elif not quoted and char == "\n":
            delimiters.append("LF")
        index += 1
    styles = set(delimiters)
    if not styles:
        return "NONE"
    if len(styles) == 1:
        return next(iter(styles))
    return "MIXED"


def _line_terminator(eol: str, baseline_raw: bytes | None = None) -> str:
    if eol == "lf":
        return "\n"
    if eol == "crlf":
        return "\r\n"
    if eol != "preserve":
        raise BridgeError(f"Unsupported output EOL mode: {eol}", 2)
    if baseline_raw is None:
        return "\n"
    record_eol = _record_eol(baseline_raw)
    if record_eol == "CRLF":
        return "\r\n"
    if record_eol in {"LF", "NONE"}:
        return "\n"
    raise BridgeError(
        f"Cannot preserve a {record_eol} record-delimiter style; select --eol crlf or lf",
        5,
    )


def serialize_csv(
    records: Sequence[Sequence[str]],
    *,
    eol: str = "lf",
    baseline_raw: bytes | None = None,
) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.writer(
        stream,
        delimiter=",",
        quotechar='"',
        doublequote=True,
        quoting=csv.QUOTE_MINIMAL,
        lineterminator=_line_terminator(eol, baseline_raw),
    )
    for row in records:
        writer.writerow([str(value) for value in row])
    return stream.getvalue().encode("utf-8")


def validate_canonical_bytes(raw: bytes) -> list[str]:
    errors: list[str] = []
    if raw.startswith(UTF8_BOM):
        errors.append("UTF8_BOM_PRESENT")
    try:
        raw.decode("utf-8", errors="strict")
    except UnicodeDecodeError:
        errors.append("INVALID_UTF8")
    if b"\x00" in raw:
        errors.append("NUL_PRESENT")
    return errors


def _atomic_write_new(path: Path, data: bytes) -> None:
    if path.exists():
        raise BridgeError(f"Output already exists: {path}", 7)
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(fd, "wb") as handle:
            handle.write(data)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_name, path)
    except Exception:
        try:
            os.unlink(temp_name)
        except FileNotFoundError:
            pass
        raise


def write_csv_records(
    path: Path,
    records: Sequence[Sequence[str]],
    *,
    eol: str = "lf",
    baseline_raw: bytes | None = None,
) -> bytes:
    raw = serialize_csv(records, eol=eol, baseline_raw=baseline_raw)
    errors = validate_canonical_bytes(raw)
    if errors:
        raise BridgeError("Generated CSV violates Canonical policy: " + ",".join(errors), 7)
    _atomic_write_new(path, raw)
    return raw


def write_manifest(path: Path | None, payload: dict[str, Any]) -> None:
    if path is None:
        return
    if path.exists():
        raise BridgeError(f"Manifest output already exists: {path}", 7)
    data = (json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n").encode(
        "utf-8"
    )
    _atomic_write_new(path, data)


def _base_manifest(command: str, doc: CsvDocument | None = None) -> dict[str, Any]:
    result: dict[str, Any] = {
        "tool": "csv_excel_bridge.py",
        "format_version": FORMAT_VERSION,
        "command": command,
        "result": "PASS",
        "input_path": str(doc.path) if doc else "",
        "input_sha256": sha256_bytes(doc.raw) if doc else "",
        "output_path": "",
        "output_sha256": "",
        "record_count": doc.record_count if doc else 0,
        "column_count": doc.column_count if doc else 0,
        "encoding": doc.encoding if doc else "",
        "bom": doc.bom if doc else False,
        "eol": _eol_label(doc.raw) if doc else "",
        "warnings": [],
        "errors": [],
    }
    return result


def parse_name_list(value: str | None) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in value.split(",") if item.strip()]


def _date_formats(extra: Sequence[str] | None = None) -> tuple[str, ...]:
    return tuple(dict.fromkeys((*DEFAULT_DATE_FORMATS, *(extra or ()))))


def normalize_date_text(value: str, formats: Sequence[str] | None = None) -> str:
    if value == "":
        return ""
    matches: set[str] = set()
    for fmt in _date_formats(formats):
        try:
            parsed = datetime.strptime(value, fmt).date()
        except ValueError:
            continue
        matches.add(parsed.isoformat())
    if len(matches) != 1:
        raise BridgeError("Unsupported or ambiguous date lexical value", 5)
    return next(iter(matches))


def _resolve_date_columns(header: Sequence[str], requested: Sequence[str]) -> set[int]:
    duplicates = {name for name in requested if requested.count(name) > 1}
    if duplicates:
        raise BridgeError("Duplicate declared date column name", 2)
    indexes: set[int] = set()
    for name in requested:
        hits = [index for index, header_name in enumerate(header) if header_name == name]
        if len(hits) != 1:
            raise BridgeError(f"Declared date column must match exactly one header: {name}", 5)
        indexes.add(hits[0])
    return indexes


def _normalized_records(
    records: Sequence[Sequence[str]], date_columns: Sequence[str], date_formats: Sequence[str]
) -> tuple[list[list[str]], int]:
    if not records:
        if date_columns:
            raise BridgeError("Cannot resolve date columns in an empty CSV", 5)
        return [], 0
    indexes = _resolve_date_columns(records[0], date_columns)
    result: list[list[str]] = [list(records[0])]
    normalized = 0
    for row in records[1:]:
        converted = list(row)
        for index in indexes:
            if index >= len(converted):
                raise BridgeError("Declared date column is missing from a record", 4)
            old = converted[index]
            new = normalize_date_text(old, date_formats)
            converted[index] = new
            if new != old:
                normalized += 1
        result.append(converted)
    return result, normalized


def _rectangle_intersects_merged(ws: Any, rows: int, columns: int) -> str | None:
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= rows and merged.min_col <= columns:
            if merged.max_row >= 1 and merged.max_col >= 1:
                return str(merged)
    return None


def _copy_visual_style(source: Cell, target: Cell) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def export_csv_to_xlsx(
    input_path: Path,
    output_path: Path,
    *,
    encoding: str = "auto-utf8",
    sheet_name: str = "CSV",
    template: Path | None = None,
    freeze_header: bool = True,
    autofilter: bool = True,
    manifest_path: Path | None = None,
    baseline_sha256: str | None = None,
    date_columns: Sequence[str] = (),
    date_formats: Sequence[str] = (),
) -> dict[str, Any]:
    doc = read_csv_records(input_path, encoding)
    records, date_count = _normalized_records(doc.records, date_columns, date_formats)
    rows = len(records)
    columns = max((len(row) for row in records), default=0)
    if output_path.resolve() == input_path.resolve() or (
        template and output_path.resolve() == template.resolve()
    ):
        raise BridgeError("Output must not overwrite an input or template", 7)
    if output_path.exists():
        raise BridgeError(f"Output already exists: {output_path}", 7)

    formatting_warnings: list[str] = []
    if template:
        try:
            wb = load_workbook(template, data_only=False)
        except Exception as exc:
            raise BridgeError(f"Template workbook cannot be opened: {type(exc).__name__}", 5) from exc
        if sheet_name not in wb.sheetnames:
            raise BridgeError(f"Requested data worksheet does not exist: {sheet_name}", 5)
        ws = wb[sheet_name]
        merged = _rectangle_intersects_merged(ws, rows, columns)
        if merged:
            raise BridgeError(f"Merged range blocks CSV data rectangle: {merged}", 5)
        old_max_row, old_max_column = ws.max_row, ws.max_column
        for row in ws.iter_rows(
            min_row=1,
            max_row=max(old_max_row, rows),
            min_col=1,
            max_col=max(old_max_column, columns),
        ):
            for cell in row:
                cell.value = None
        styled_rows = min(old_max_row, rows)
        styled_columns = min(old_max_column, columns)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        old_max_row = old_max_column = 0
        styled_rows = styled_columns = 0

    for row_index, record in enumerate(records, 1):
        for column_index in range(1, columns + 1):
            value = record[column_index - 1] if column_index <= len(record) else ""
            cell = ws.cell(row=row_index, column=column_index)
            cell.value = value if value != "" else None
            cell.number_format = "@"
            if template is None:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    if template is None and records:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="4472C4")
        for column_index in range(1, columns + 1):
            longest = max(
                (len(records[row][column_index - 1]) for row in range(rows) if column_index <= len(records[row])),
                default=0,
            )
            ws.column_dimensions[get_column_letter(column_index)].width = min(max(longest + 2, 10), 60)

    if freeze_header and rows:
        ws.freeze_panes = "A2"
    if autofilter and rows and columns:
        ws.auto_filter.ref = f"A1:{get_column_letter(columns)}{rows}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary = tempfile.mkstemp(prefix=f".{output_path.name}.", suffix=".xlsx", dir=output_path.parent)
    os.close(fd)
    try:
        wb.save(temporary)
        os.replace(temporary, output_path)
    except Exception:
        try:
            os.unlink(temporary)
        except FileNotFoundError:
            pass
        raise

    manifest = _base_manifest("export", doc)
    manifest.update(
        {
            "output_path": str(output_path),
            "output_sha256": sha256_file(output_path),
            "baseline_sha256": baseline_sha256 or "",
            "date_columns": list(date_columns),
            "date_cells_normalized": date_count,
            "date_normalization_errors": 0,
            "template_path": str(template) if template else "",
            "template_sha256": sha256_file(template) if template else "",
            "formatting_recovery": bool(template),
            "styled_rows_reused": styled_rows,
            "styled_columns_reused": styled_columns,
            "new_rows_default_styled": max(rows - old_max_row, 0),
            "new_columns_default_styled": max(columns - old_max_column, 0),
            "formatting_warnings": formatting_warnings,
        }
    )
    write_manifest(manifest_path, manifest)
    return manifest


def _last_value_cell(ws: Any) -> tuple[int, int]:
    max_row = max_column = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_row = max(max_row, cell.row)
                max_column = max(max_column, cell.column)
    return max_row, max_column


def _cell_to_text(
    cell: Cell,
    *,
    is_date_column: bool,
    date_formats: Sequence[str],
) -> tuple[str, bool]:
    if cell.data_type == "f":
        raise BridgeError(f"Formula cell rejected at {cell.coordinate}", 5)
    if cell.data_type == "e":
        raise BridgeError(f"Error cell rejected at {cell.coordinate}", 5)
    value = cell.value
    if value is None:
        return "", False
    if isinstance(value, bool):
        raise BridgeError(f"Boolean cell rejected at {cell.coordinate}", 5)
    if isinstance(value, datetime):
        if not is_date_column:
            raise BridgeError(f"Date/datetime cell outside declared date column at {cell.coordinate}", 5)
        if value.time() != time(0, 0, 0):
            raise BridgeError(f"Datetime has non-zero time component at {cell.coordinate}", 5)
        return value.date().isoformat(), True
    if isinstance(value, date):
        if not is_date_column:
            raise BridgeError(f"Date cell outside declared date column at {cell.coordinate}", 5)
        return value.isoformat(), True
    if isinstance(value, (int, float, complex)):
        raise BridgeError(f"Numeric cell rejected at {cell.coordinate}", 5)
    if not isinstance(value, str):
        raise BridgeError(f"Unsupported cell type at {cell.coordinate}", 5)
    if is_date_column:
        normalized = normalize_date_text(value, date_formats)
        return normalized, normalized != value
    return value, False


def import_xlsx_to_csv(
    input_path: Path,
    output_path: Path,
    *,
    baseline: Path | None = None,
    sheet_name: str = "CSV",
    allow_header_change: bool = False,
    allow_row_count_change: bool = False,
    allow_column_count_change: bool = False,
    manifest_path: Path | None = None,
    date_columns: Sequence[str] = (),
    date_formats: Sequence[str] = (),
    output_eol: str = "preserve",
) -> dict[str, Any]:
    if output_path.exists():
        raise BridgeError(f"Output already exists: {output_path}", 7)
    try:
        wb = load_workbook(input_path, data_only=False)
    except Exception as exc:
        raise BridgeError(f"Workbook cannot be opened: {type(exc).__name__}", 5) from exc
    unexpected = [
        ws.title
        for ws in wb.worksheets
        if ws.title not in {sheet_name, "__UADC_BRIDGE__"} and ws.sheet_state == "visible"
    ]
    if unexpected or sheet_name not in wb.sheetnames:
        raise BridgeError("Unexpected or missing worksheet", 5)
    ws = wb[sheet_name]
    baseline_doc = read_csv_records(baseline, "auto-utf8") if baseline else None
    actual_rows, actual_columns = _last_value_cell(ws)
    if baseline_doc:
        expected_rows = baseline_doc.record_count
        expected_columns = baseline_doc.column_count
        rows = max(expected_rows, actual_rows) if allow_row_count_change else expected_rows
        columns = max(expected_columns, actual_columns) if allow_column_count_change else expected_columns
        missing_rows_have_values = any(
            any(value != "" for value in row)
            for row in baseline_doc.records[actual_rows:]
        )
        missing_columns_have_values = any(
            any(value != "" for value in row[actual_columns:])
            for row in baseline_doc.records
        )
        if not allow_row_count_change and (
            actual_rows > expected_rows or (actual_rows < expected_rows and missing_rows_have_values)
        ):
            raise BridgeError("Row count differs from baseline", 6)
        if not allow_column_count_change and (
            actual_columns > expected_columns
            or (actual_columns < expected_columns and missing_columns_have_values)
        ):
            raise BridgeError("Column count differs from baseline", 6)
    else:
        rows, columns = actual_rows, actual_columns

    header = ["" if ws.cell(1, col).value is None else str(ws.cell(1, col).value) for col in range(1, columns + 1)] if rows else []
    date_indexes = _resolve_date_columns(header, date_columns) if rows else set()
    records: list[list[str]] = []
    date_count = 0
    for row_index in range(1, rows + 1):
        record: list[str] = []
        for column_index in range(1, columns + 1):
            text, normalized = _cell_to_text(
                ws.cell(row_index, column_index),
                is_date_column=(row_index > 1 and column_index - 1 in date_indexes),
                date_formats=date_formats,
            )
            record.append(text)
            date_count += int(normalized)
        records.append(record)

    header_changed = False
    row_count_changed = False
    column_count_changed = False
    changed_cells: list[dict[str, Any]] = []
    if baseline_doc:
        header_changed = bool(records and baseline_doc.records and records[0] != baseline_doc.records[0])
        row_count_changed = len(records) != len(baseline_doc.records)
        column_count_changed = (
            max((len(row) for row in records), default=0) != baseline_doc.column_count
        )
        if header_changed and not allow_header_change:
            raise BridgeError("Header differs from baseline", 6)
        if row_count_changed and not allow_row_count_change:
            raise BridgeError("Row count differs from baseline", 6)
        if column_count_changed and not allow_column_count_change:
            raise BridgeError("Column count differs from baseline", 6)
        for row_index in range(min(len(records), len(baseline_doc.records))):
            for column_index in range(
                min(len(records[row_index]), len(baseline_doc.records[row_index]))
            ):
                old = baseline_doc.records[row_index][column_index]
                new = records[row_index][column_index]
                if old != new:
                    changed_cells.append(
                        {
                            "row": row_index + 1,
                            "column": column_index + 1,
                            "old_value_sha256": value_sha256(old),
                            "new_value_sha256": value_sha256(new),
                        }
                    )

    raw = write_csv_records(
        output_path,
        records,
        eol=output_eol,
        baseline_raw=baseline_doc.raw if baseline_doc else None,
    )
    manifest = {
        "tool": "csv_excel_bridge.py",
        "format_version": FORMAT_VERSION,
        "command": "import",
        "result": "PASS",
        "input_path": str(input_path),
        "input_sha256": sha256_file(input_path),
        "output_path": str(output_path),
        "output_sha256": sha256_bytes(raw),
        "record_count": len(records),
        "column_count": max((len(row) for row in records), default=0),
        "encoding": "utf-8",
        "bom": False,
        "eol": _eol_label(raw),
        "record_eol": _record_eol(raw),
        "requested_output_eol": output_eol,
        "warnings": [],
        "errors": [],
        "baseline_path": str(baseline) if baseline else "",
        "baseline_sha256": sha256_file(baseline) if baseline else "",
        "header_changed": header_changed,
        "row_count_changed": row_count_changed,
        "column_count_changed": column_count_changed,
        "changed_cell_count": len(changed_cells),
        "changed_cells": changed_cells,
        "date_columns": list(date_columns),
        "date_cells_normalized": date_count,
        "date_normalization_errors": 0,
    }
    write_manifest(manifest_path, manifest)
    return manifest


def compare_csv_logical(baseline: Path, candidate: Path) -> tuple[str, dict[str, Any]]:
    try:
        left = read_csv_records(baseline, "auto-utf8")
        right = read_csv_records(candidate, "auto-utf8")
    except BridgeError as exc:
        return "PARSE_ERROR", {"result": "PARSE_ERROR", "errors": [str(exc)]}
    normalized_left = [
        [value.replace("\r\n", "\n").replace("\r", "\n") for value in row]
        for row in left.records
    ]
    normalized_right = [
        [value.replace("\r\n", "\n").replace("\r", "\n") for value in row]
        for row in right.records
    ]
    if left.raw == right.raw:
        result = "BYTE_IDENTICAL"
    elif left.records == right.records or normalized_left == normalized_right:
        result = "LOGICALLY_IDENTICAL_BYTES_DIFFER"
    else:
        result = "LOGICAL_DIFFERENCE"
    return result, {
        "tool": "csv_excel_bridge.py",
        "format_version": FORMAT_VERSION,
        "command": "compare",
        "result": result,
        "baseline_path": str(baseline),
        "baseline_sha256": sha256_bytes(left.raw),
        "candidate_path": str(candidate),
        "candidate_sha256": sha256_bytes(right.raw),
        "baseline_record_count": left.record_count,
        "candidate_record_count": right.record_count,
    }


def validate_csv(
    path: Path, *, canonical: bool = False, expected_columns: int | None = None
) -> dict[str, Any]:
    doc = read_csv_records(path, "auto-utf8")
    errors = validate_canonical_bytes(doc.raw) if canonical else []
    widths = {len(row) for row in doc.records}
    if expected_columns is not None and widths != {expected_columns}:
        errors.append("UNEXPECTED_COLUMN_COUNT")
    result = _base_manifest("validate", doc)
    result["errors"] = errors
    result["result"] = "PASS" if not errors else "FAIL"
    if errors:
        raise BridgeError("CSV validation failed: " + ",".join(errors), 3)
    return result


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)

    export = sub.add_parser("export")
    export.add_argument("input", type=Path)
    export.add_argument("-o", "--output", type=Path, required=True)
    export.add_argument("--encoding", choices=("utf-8", "utf-8-sig", "auto-utf8"), default="auto-utf8")
    export.add_argument("--sheet-name", default="CSV")
    export.add_argument("--template", type=Path)
    export.add_argument("--freeze-header", action=argparse.BooleanOptionalAction, default=True)
    export.add_argument("--autofilter", action=argparse.BooleanOptionalAction, default=True)
    export.add_argument("--manifest", type=Path)
    export.add_argument("--baseline-sha256")
    export.add_argument("--date-columns")
    export.add_argument("--date-input-formats")

    imported = sub.add_parser("import")
    imported.add_argument("input", type=Path)
    imported.add_argument("-o", "--output", type=Path, required=True)
    imported.add_argument("--baseline", type=Path)
    imported.add_argument("--sheet-name", default="CSV")
    imported.add_argument("--allow-header-change", action="store_true")
    imported.add_argument("--allow-row-count-change", action="store_true")
    imported.add_argument("--allow-column-count-change", action="store_true")
    imported.add_argument("--manifest", type=Path)
    imported.add_argument("--date-columns")
    imported.add_argument("--date-input-formats")
    imported.add_argument("--eol", choices=("preserve", "crlf", "lf"), default="preserve")

    validate = sub.add_parser("validate")
    validate.add_argument("input", type=Path)
    validate.add_argument("--canonical", action="store_true")
    validate.add_argument("--expected-columns", type=int)
    validate.add_argument("--manifest", type=Path)

    compare = sub.add_parser("compare")
    compare.add_argument("baseline", type=Path)
    compare.add_argument("candidate", type=Path)
    compare.add_argument("--manifest", type=Path)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        if args.command == "export":
            result = export_csv_to_xlsx(
                args.input,
                args.output,
                encoding=args.encoding,
                sheet_name=args.sheet_name,
                template=args.template,
                freeze_header=args.freeze_header,
                autofilter=args.autofilter,
                manifest_path=args.manifest,
                baseline_sha256=args.baseline_sha256,
                date_columns=parse_name_list(args.date_columns),
                date_formats=parse_name_list(args.date_input_formats),
            )
        elif args.command == "import":
            result = import_xlsx_to_csv(
                args.input,
                args.output,
                baseline=args.baseline,
                sheet_name=args.sheet_name,
                allow_header_change=args.allow_header_change,
                allow_row_count_change=args.allow_row_count_change,
                allow_column_count_change=args.allow_column_count_change,
                manifest_path=args.manifest,
                date_columns=parse_name_list(args.date_columns),
                date_formats=parse_name_list(args.date_input_formats),
                output_eol=args.eol,
            )
        elif args.command == "validate":
            result = validate_csv(
                args.input, canonical=args.canonical, expected_columns=args.expected_columns
            )
            write_manifest(args.manifest, result)
        else:
            status, result = compare_csv_logical(args.baseline, args.candidate)
            write_manifest(args.manifest, result)
            print(status)
            return 0 if status != "LOGICAL_DIFFERENCE" and status != "PARSE_ERROR" else 8
        print(json.dumps({"result": result["result"]}, sort_keys=True))
        return 0
    except BridgeError as exc:
        print(f"ERROR: {exc}")
        return exc.code


if __name__ == "__main__":
    raise SystemExit(main())
